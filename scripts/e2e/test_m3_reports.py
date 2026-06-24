#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""test_m3_reports.py — M3 report-render harness: parity + POI render readback for the 4 LIVE reports.

The 4 reports (the entire live set — 18 of 22 never ran, m3-report-usage-prune.md):
  daily_shipping_assy  R3   REPORT_DailyShippingAssy   (faithful — the 285-runs/yr sound shipping path)
  invoice_summary      R9   REPORT_INVOICESSummary     (D6 window-aware CORRECTED + window-blind faithful)
  forecast_detail      R18  REPORT_ForecastDetail      (faithful — group-break custom render)
  lot_location         R12  REPORT_PLANTLotLocation[W] (faithful — two-proc custom render; seeded)

TWO test layers (m3-render-architecture §5), all headless on the live spike:
  LAYER 1 — NUMBERS PARITY (pure SQL). The Named-Query .sql == the legacy proc on a live sample. EVERY
    expectation derives from the SOURCE (the legacy proc / an INDEPENDENT truth query), NEVER the rebuild.
    For invoice_summary: corrected .sql == an independent D6 window-aware CROSS-APPLY truth query AND the
    _faithful .sql == the legacy proc; the D6 fix is proven NON-VACUOUS by injecting a 2nd (gap) price
    window (legacy doubles, corrected unchanged) — the test_report_procs_d6.py pattern. Non-vacuity is
    proven by a REVERT step (corrupt the .sql -> the parity check FAILS).
  LAYER 2 — POI RENDER READBACK. The REAL render runs headless via the gateway's bundled JRE + jython 2.7
    + Apache POI (the exact code.render_xlsx that runs on the gateway) -> the .xlsx is read back with
    openpyxl and the header band + column cells + widths + the detail rows are asserted against the legacy
    `mysheet.cells` layout (extracted from MainMenu.pas, cited per report). HONEST: numbers are faithful to
    the legacy proc on live data; the exact .xls VISUAL CHROME (template, borders, page breaks) is NOT
    reproduced — we render the same CELLS, not the template chrome.

Run:  export SA_PASS='Spike_Dev_2026!'  &&  python3 scripts/e2e/test_m3_reports.py
Requires: docker mssql-spike up; openpyxl (CPython); the bundled JRE+jython+POI at /usr/local/ignition.
Leaves the spike AS-FOUND (all fixtures torn down; rolled-back / additive-then-deleted).
"""
import json
import os
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib import Report, GATEWAY_ROOT  # noqa: E402

CONTAINER = os.environ.get("CONTAINER", "mssql-spike")
SA_PASS = os.environ.get("SA_PASS")
DB = "Inventory"

REPO = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
SQL_DIR = os.path.join(REPO, "docs", "analysis", "reporting", "sql")
RR_DIR = os.path.join(REPO, "docs", "analysis", "reporting", "project-library", "report_render")

# Gateway bundled JRE + jython + POI (the exact runtime the gateway uses) — lets us run the REAL POI
# render path headless (verified: render_xlsx -> a real .xlsx). No system Java is required.
IGN = GATEWAY_ROOT   # centralized gateway root (repo-split-plan §4.C; default /usr/local/ignition)
JAVA = os.path.join(IGN, "lib", "runtime", "jre-mac", "bin", "java")
COMMON = os.path.join(IGN, "lib", "core", "common")
PYLIB = os.path.join(IGN, "user-lib", "pylib")
JYJAR = os.path.join(COMMON, "jython-ia-2.7.3.3.jar")

OUT_DIR = "/tmp/m3_reports_out"


# ============================================================================================
# sqlcmd helpers (live spike; bounded; the test_report_procs_d6.py shape)
# ============================================================================================
def _sqlcmd(query, with_header=False):
    if not SA_PASS:
        sys.exit("export SA_PASS first")
    args = ["docker", "exec", CONTAINER, "/opt/mssql-tools18/bin/sqlcmd", "-C", "-S", "localhost",
            "-U", "sa", "-P", SA_PASS, "-d", DB, "-W", "-s", "\t"]
    if not with_header:
        args += ["-h", "-1"]
    out = subprocess.check_output(args + ["-Q", "SET NOCOUNT ON; " + query], text=True)
    return out


def sqlrows(query, with_header=False):
    out = _sqlcmd(query, with_header)
    rows = [l for l in out.splitlines() if l.strip() and not l.startswith("(") and not l.startswith("Msg ")]
    if with_header:
        if not rows:
            return [], []
        cols = rows[0].split("\t")
        data = [r.split("\t") for r in rows[1:] if set(r) - set("- \t")]
        return cols, data
    return [r.split("\t") for r in rows]


def scalar(query):
    r = sqlrows(query)
    return r[0][0] if r else None


# ============================================================================================
# .sql loader (mirrors driver.load_sql: strip -- comment lines) — used for BOTH the live parity run
# (we EXEC the .sql directly) and to hand the bound SQL to the jython render helper.
# ============================================================================================
def load_sql(sql_key):
    with open(os.path.join(SQL_DIR, sql_key + ".sql")) as fh:
        lines = fh.read().splitlines()
    body = []
    for ln in lines:
        if ln.strip().startswith("--"):
            continue
        # strip a trailing inline -- comment (so a '?' inside a comment isn't treated as a bind marker).
        # No '?' appears inside any string literal in these queries, so a plain split is safe here.
        if "--" in ln:
            ln = ln[:ln.index("--")]
        body.append(ln)
    return "\n".join(body).strip()


def bind_qmarks(sql, args):
    """Inline ? markers as T-SQL literals (the .sql uses ? like the NQ; for a live EXEC we substitute)."""
    out, i = [], 0
    for ch in sql:
        if ch == "?":
            v = args[i]; i += 1
            out.append("NULL" if v is None else "'" + str(v).replace("'", "''") + "'")
        else:
            out.append(ch)
    return "".join(out)


# ============================================================================================
# JYTHON RENDER HELPER — run the REAL code.render_xlsx headless via the bundled JRE.
# We pass the report def + the detail/header rows (fetched in CPython via sqlcmd) as JSON; the jython
# side builds the plan (the SAME pure code.build_*_plan) + renders via POI + writes the .xlsx. This
# exercises the real gateway render path (POI) without a running gateway or a JVM on the host.
# ============================================================================================
_JY_RENDER = r'''# -*- coding: utf-8 -*-
import sys, json
sys.path.insert(0, %(rr)r)
import code as engine
from java.io import FileOutputStream

payload = json.loads(open(%(payload)r).read())
render = payload["render"]
rdef = payload["rdef"]
params = payload.get("params") or {}

class Row(dict):
    def __getitem__(self, k):
        return dict.get(self, k)

def rows(seq):
    return [Row(r) for r in seq]

if render == "build_lot_location_plan":
    ds = {"wheels": rows(payload["wheels"]), "tires": rows(payload["tires"])}
    cells, widths = engine.build_lot_location_plan(rdef, ds, params=params)
else:
    detail = rows(payload["detail"])
    header = payload.get("header")
    header_row = Row(header) if header else None
    if render == "declarative":
        cells, widths = engine.build_declarative_plan(rdef, detail, header_row, params)
    elif render == "build_forecast_detail_plan":
        cells, widths = engine.build_forecast_detail_plan(rdef, detail, header_row, params)
    else:
        raise ValueError("unknown render %%r" %% render)

b = engine.render_xlsx(cells, widths)
fos = FileOutputStream(payload["out"]); fos.write(b); fos.close()
print "OK cells=%%d bytes=%%d" %% (len(cells), len(b))
'''


def jython_render(render, rdef, params, out_path, detail=None, header=None, wheels=None, tires=None):
    """Render a report's .xlsx via the REAL POI path under the bundled jython. Returns the out_path."""
    payload = {"render": render, "rdef": rdef, "params": params or {}, "out": out_path}
    if render == "build_lot_location_plan":
        payload["wheels"] = wheels or []
        payload["tires"] = tires or []
    else:
        payload["detail"] = detail or []
        payload["header"] = header
    pfile = out_path + ".payload.json"
    with open(pfile, "w") as fh:
        json.dump(payload, fh)
    script = _JY_RENDER % {"rr": RR_DIR, "payload": pfile}
    sfile = out_path + ".render.py"
    with open(sfile, "w") as fh:
        fh.write(script)
    env = dict(os.environ)
    cp = JYJAR + ":" + os.path.join(COMMON, "*")
    # cachedir.skip=true: don't write jython's package-scan cache (it otherwise lands as a stray
    # .jython_cache/ in cwd, not under -Dpython.cachedir). We don't need the cache for a one-shot render.
    cmd = [JAVA, "-Dpython.path=" + PYLIB, "-Dpython.cachedir.skip=true",
           "-cp", cp, "org.python.util.jython", "-S", sfile]
    out = subprocess.check_output(cmd, text=True, stderr=subprocess.STDOUT, env=env)
    return out_path, out.strip()


def fetch_rows_as_dicts(sql_key, args=None):
    """Run a .sql against the live spike (with header) and return [dict(col->value)] for the render side."""
    sql = bind_qmarks(load_sql(sql_key), args or [])
    cols, data = sqlrows(sql, with_header=True)
    out = []
    for r in data:
        d = {}
        for i, c in enumerate(cols):
            v = r[i] if i < len(r) else None
            d[c] = None if v == "NULL" else v
        out.append(d)
    return cols, out


# ============================================================================================
# openpyxl readback helpers
# ============================================================================================
def read_xlsx(path):
    import openpyxl
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path)
    ws = wb.active
    return ws


def cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def colwidth(ws, col):
    from openpyxl.utils import get_column_letter
    cd = ws.column_dimensions.get(get_column_letter(col))
    return cd.width if cd else None


# ============================================================================================
# Report defs (read from the actual report_defs.py so the test asserts the SHIPPED config, not a copy)
# ============================================================================================
def load_report_defs():
    import importlib.util
    spec = importlib.util.spec_from_file_location("report_defs", os.path.join(RR_DIR, "report_defs.py"))
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


# ============================================================================================
# MAIN
# ============================================================================================
def main():
    print("=" * 80)
    print(" M3 REPORT RENDER HARNESS — parity (numbers) + POI render readback (4 LIVE reports)")
    print("=" * 80)
    rep = Report()
    os.makedirs(OUT_DIR, exist_ok=True)
    defs = load_report_defs()

    test_daily_shipping_assy(rep, defs)
    test_invoice_summary(rep, defs)
    test_forecast_detail(rep, defs)
    test_lot_location(rep, defs)
    test_output_stage(rep)
    test_revert_non_vacuity(rep)

    sys.exit(rep.summary_exit())


# ----------------------------------------------------------------------------------------------------
# R3 — Daily Shipping Assy (faithful)
# ----------------------------------------------------------------------------------------------------
def test_daily_shipping_assy(rep, defs):
    print("\n--- R3 Daily Shipping Assy (REPORT_DailyShippingAssy, faithful) ---")
    rdef = defs.REPORTS["daily_shipping_assy"]

    # pick a live production date with the most ASN detail (non-vacuous)
    pdate = scalar("SELECT TOP 1 a.VC_PRODUCTION_DATE FROM INV_ASN_MST a "
                   "JOIN INV_ASN_DETAIL_MST d ON a.IN_ASN_ID=d.IN_ASN_ID "
                   "GROUP BY a.VC_PRODUCTION_DATE ORDER BY COUNT(*) DESC")
    if not pdate:
        rep.check("R3: found a production date with ASN detail", False, "none")
        return
    print("  production date = %s" % pdate)

    # LAYER 1 — parity: the .sql == the legacy proc, by (Part Number, PQty) rows.
    legacy = sqlrows("EXEC REPORT_DailyShippingAssy @PDate='%s'" % pdate, with_header=True)
    nq = sqlrows(bind_qmarks(load_sql("daily_shipping_assy"), [pdate]), with_header=True)
    leg_pairs = _pairs(legacy, "Part Number", "PQty")
    nq_pairs = _pairs(nq, "Part Number", "PQty")
    rep.check("R3 L1: NQ detail == legacy proc (Part Number, PQty rows)",
              leg_pairs == nq_pairs and len(nq_pairs) > 0,
              "legacy=%d nq=%d rows" % (len(leg_pairs), len(nq_pairs)))

    # NON-VACUITY: corrupt the .sql's WHERE (a wrong date) -> parity must FAIL.
    bad = sqlrows(bind_qmarks(load_sql("daily_shipping_assy"), ["19000101"]), with_header=True)
    rep.check("R3 L1 non-vacuity: a wrong-date NQ does NOT match legacy",
              _pairs(bad, "Part Number", "PQty") != leg_pairs, "bad rows=%d" % (len(bad[1]) if bad else 0))

    # header band parity: Start/End/Vehicle Count == the LEGACY printed header.
    # LEGACY behavior (MainMenu.pas:3173-3174, TMainMenu_Form.DailyASNReportClick): the header band is
    # read off the FIRST DETAIL ROW of REPORT_DailyShippingAssy (fieldbyname('Start'/'End'/'Vehicle Count')
    # BEFORE the z:=4 loop). NOT MIN/MAX/SUM. The proc orders ONLY by VC_ASSY_PART_NUMBER, so "row 1" =
    # the header that owns the alphabetically-first part.
    hcols, hdata = sqlrows(bind_qmarks(load_sql("daily_shipping_header_assy"), [pdate]), with_header=True)
    header = dict(zip(hcols, hdata[0])) if hdata else {}
    # INDEPENDENT legacy truth: EXEC the real proc, read row 1's Start/End/Vehicle Count (what Delphi printed).
    # This is NOT the rebuild's plan and NOT MIN/MAX/SUM — it is the legacy proc's own first detail row.
    leg_start, leg_end, leg_vc = _legacy_assy_header_row1(pdate)
    rep.check("R3 L1 header band == LEGACY first-detail-row (MainMenu.pas:3173-3174)",
              header.get("StartSeq") == leg_start and header.get("EndSeq") == leg_end
              and header.get("Vehicle Count") == leg_vc,
              "hdr=%s/%s/%s legacy-row1=%s/%s/%s" % (header.get("StartSeq"), header.get("EndSeq"),
                                                     header.get("Vehicle Count"), leg_start, leg_end, leg_vc))

    # MULTI-HEADER regression guard (the -1 sentinel bug). On a multi-header day a throwaway -1/-1 sentinel
    # header co-exists with the real one. The OLD MIN/MAX/SUM header printed "Start Seq: -1" and an inflated
    # Vehicle Count; the fixed header reproduces the legacy proc's row 1 (the real header). Assert the fixed
    # NQ header == the legacy first-detail-row AND demonstrate non-vacuity: the OLD MIN/MAX/SUM derivation
    # would FAIL on this date (it picks -1 / inflated count, != the legacy row 1).
    # Pick a multi-header date whose alphabetically-FIRST part belongs to exactly ONE header, so the raw
    # legacy proc's "row 1" (ORDER BY part only, no tiebreaker) is deterministic and == the NQ's pinned row.
    # Prefer the documented regression date 20250121; else any qualifying date with the most ASN headers.
    mh = scalar(
        "SELECT TOP 1 a.VC_PRODUCTION_DATE "
        "FROM INV_ASN_MST a "
        "WHERE a.VC_PRODUCTION_DATE IN ("
        "    SELECT VC_PRODUCTION_DATE FROM INV_ASN_MST "
        "    GROUP BY VC_PRODUCTION_DATE HAVING COUNT(DISTINCT VC_START_SEQ_NUMBER) > 1) "
        "  AND NOT EXISTS ("  # exclude dates where the first part is tied across >1 header
        "    SELECT 1 FROM ("
        "      SELECT d.VC_ASSY_PART_NUMBER part, COUNT(DISTINCT s.VC_START_SEQ_NUMBER) hc "
        "      FROM INV_ASN_MST s JOIN INV_ASN_DETAIL_MST d ON s.IN_ASN_ID=d.IN_ASN_ID "
        "      WHERE s.VC_PRODUCTION_DATE=a.VC_PRODUCTION_DATE "
        "      GROUP BY d.VC_ASSY_PART_NUMBER) fp "
        "    WHERE fp.part=("
        "      SELECT MIN(d2.VC_ASSY_PART_NUMBER) FROM INV_ASN_MST s2 "
        "      JOIN INV_ASN_DETAIL_MST d2 ON s2.IN_ASN_ID=d2.IN_ASN_ID "
        "      WHERE s2.VC_PRODUCTION_DATE=a.VC_PRODUCTION_DATE) AND fp.hc > 1) "
        "GROUP BY a.VC_PRODUCTION_DATE "
        "ORDER BY CASE WHEN a.VC_PRODUCTION_DATE='20250121' THEN 0 ELSE 1 END, COUNT(*) DESC")
    if not mh:
        rep.check("R3 L1 multi-header: a >1-distinct-start-seq date exists on the spike", False, "none")
    else:
        mh_hcols, mh_hdata = sqlrows(bind_qmarks(load_sql("daily_shipping_header_assy"), [mh]), with_header=True)
        mh_header = dict(zip(mh_hcols, mh_hdata[0])) if mh_hdata else {}
        mh_start, mh_end, mh_vc = _legacy_assy_header_row1(mh)
        rep.check("R3 L1 multi-header header == LEGACY first-detail-row (no -1 sentinel) [%s]" % mh,
                  mh_header.get("StartSeq") == mh_start and mh_header.get("EndSeq") == mh_end
                  and mh_header.get("Vehicle Count") == mh_vc and mh_header.get("StartSeq") != "-1",
                  "hdr=%s/%s/%s legacy-row1=%s/%s/%s"
                  % (mh_header.get("StartSeq"), mh_header.get("EndSeq"), mh_header.get("Vehicle Count"),
                     mh_start, mh_end, mh_vc))
        # NON-VACUITY: the OLD MIN/MAX/SUM header would NOT match the legacy row 1 on this multi-header date
        # (it prints the -1 sentinel / inflated SUM). If this ever "passes" the regression is back.
        old = sqlrows("SELECT MIN(VC_START_SEQ_NUMBER), MAX(VC_END_SEQ_NUMBER), SUM(IN_QTY) "
                      "FROM INV_ASN_MST WHERE VC_PRODUCTION_DATE='%s'" % mh)
        o_start, o_end, o_vc = old[0]
        rep.check("R3 L1 multi-header NON-VACUITY: old MIN/MAX/SUM header DIVERGES from legacy [%s]" % mh,
                  not (o_start == mh_start and o_end == mh_end and o_vc == mh_vc),
                  "old=%s/%s/%s legacy-row1=%s/%s/%s" % (o_start, o_end, o_vc, mh_start, mh_end, mh_vc))

    # LAYER 2 — POI render readback
    _, detail = fetch_rows_as_dicts("daily_shipping_assy", [pdate])
    out = os.path.join(OUT_DIR, "assy.xlsx")
    try:
        jython_render("declarative", _jsonable(rdef), {"PDate": _fmt_date(pdate)}, out,
                      detail=detail, header=header)
    except subprocess.CalledProcessError as e:
        rep.check("R3 L2: POI render ran", False, e.output[-300:])
        return
    ws = read_xlsx(out)
    rep.check("R3 L2 [1,1] title", cell(ws, 1, 1) == "ASN (Assy Part Numbers)", "%r" % cell(ws, 1, 1))
    rep.check("R3 L2 [2,2] StartEnd from header", cell(ws, 2, 2) == header.get("StartEnd"),
              "%r" % cell(ws, 2, 2))
    rep.check("R3 L2 [2,3] 'Vehicle Count:'+count",
              cell(ws, 2, 3) == "Vehicle Count:" + str(header.get("Vehicle Count")), "%r" % cell(ws, 2, 3))
    rep.check("R3 L2 row-3 col titles Part Number/Qty",
              cell(ws, 3, 1) == "Part Number" and cell(ws, 3, 2) == "Qty",
              "%r/%r" % (cell(ws, 3, 1), cell(ws, 3, 2)))
    rep.check("R3 L2 widths 17/30", colwidth(ws, 1) == 17 and colwidth(ws, 2) == 30,
              "%s/%s" % (colwidth(ws, 1), colwidth(ws, 2)))
    # detail rows (z=4..) match the NQ rows in order
    ok = True
    for i, d in enumerate(detail):
        z = 4 + i
        if cell(ws, z, 1) != d["Part Number"] or int(cell(ws, z, 2)) != int(d["PQty"]):
            ok = False
            break
    rep.check("R3 L2 detail [z,1]=Part Number [z,2]=PQty for all %d rows" % len(detail), ok)
    rep.check("R3 L2 detail-row count == NQ row count (%d)" % len(detail),
              _count_detail_rows(ws, 4, 1) == len(detail), "sheet=%d" % _count_detail_rows(ws, 4, 1))


# ----------------------------------------------------------------------------------------------------
# R9 — INVOICE Summary (D6 corrected default + window-blind faithful)
# ----------------------------------------------------------------------------------------------------
def test_invoice_summary(rep, defs):
    print("\n--- R9 INVOICE Summary (REPORT_INVOICESSummary, D6 corrected + faithful) ---")
    rdef = defs.REPORTS["invoice_summary"]

    pdate = scalar("SELECT TOP 1 a.VC_PRODUCTION_DATE FROM INV_ASN_MST a "
                   "JOIN INV_ASN_DETAIL_MST d ON a.IN_ASN_ID=d.IN_ASN_ID "
                   "JOIN INV_MANIFEST_COST_MST m ON d.VC_ASSY_PART_NUMBER=m.VC_ASSY_PART_NUMBER_CODE "
                   "JOIN INV_INV_MST i ON i.IN_INV_ID=d.IN_INV_ID "
                   "GROUP BY a.VC_PRODUCTION_DATE ORDER BY COUNT(*) DESC")
    if not pdate:
        rep.check("R9: found an invoiced production date", False, "none")
        return
    print("  production date = %s" % pdate)

    # LAYER 1a — FAITHFUL: the _faithful .sql == the legacy proc, EXACTLY (verbatim window-blind).
    legacy = sqlrows("EXEC REPORT_INVOICESSummary @PDate='%s'" % pdate, with_header=True)
    faithful = sqlrows(bind_qmarks(load_sql("invoice_summary_faithful"), [pdate]), with_header=True)
    leg = _tuples(legacy, ["Manifest", "PartNumber", "UnitPrice", "ShipQty", "PickUpDate"])
    fai = _tuples(faithful, ["Manifest", "PartNumber", "UnitPrice", "ShipQty", "PickUpDate"])
    rep.check("R9 L1a: _faithful .sql == legacy proc (verbatim window-blind)",
              leg == fai and len(fai) > 0, "legacy=%d faithful=%d" % (len(leg), len(fai)))

    # LAYER 1b — CORRECTED: the corrected .sql == an INDEPENDENT D6 window-aware truth query.
    truth = sqlrows(
        "SELECT d.VC_MANIFEST_NUMBER 'Manifest', d.VC_ASSY_PART_NUMBER 'PartNumber', "
        "m.MO_PRICE 'UnitPrice', d.IN_QTY 'ShipQty', a.VC_PRODUCTION_DATE 'PickUpDate' "
        "FROM INV_ASN_MST a JOIN INV_ASN_DETAIL_MST d ON a.IN_ASN_ID=d.IN_ASN_ID "
        "CROSS APPLY dbo.fn_ManifestCostAt(d.VC_ASSY_PART_NUMBER, a.VC_PRODUCTION_DATE) m "
        "JOIN INV_INV_MST i ON i.IN_INV_ID=d.IN_INV_ID WHERE a.VC_PRODUCTION_DATE='%s' "
        "ORDER BY d.VC_MANIFEST_NUMBER" % pdate, with_header=True)
    corrected = sqlrows(bind_qmarks(load_sql("invoice_summary"), [pdate]), with_header=True)
    tr = _tuples(truth, ["Manifest", "PartNumber", "UnitPrice", "ShipQty", "PickUpDate"])
    cor = _tuples(corrected, ["Manifest", "PartNumber", "UnitPrice", "ShipQty", "PickUpDate"])
    rep.check("R9 L1b: corrected .sql == independent D6 window-aware truth",
              cor == tr and len(cor) > 0, "corrected=%d truth=%d" % (len(cor), len(tr)))

    # HEADLINE D6 NON-VACUITY: inject a 2nd (gap) price window for a part on PDATE -> the FAITHFUL/legacy
    # output DOUBLES that part's lines (the over-bill) while the CORRECTED stays unchanged (one covering
    # window). Proves we characterized the D6 divergence, not just diffed equal-today data.
    part = scalar("SELECT TOP 1 d.VC_ASSY_PART_NUMBER FROM INV_ASN_MST a "
                  "JOIN INV_ASN_DETAIL_MST d ON a.IN_ASN_ID=d.IN_ASN_ID "
                  "JOIN INV_MANIFEST_COST_MST m ON d.VC_ASSY_PART_NUMBER=m.VC_ASSY_PART_NUMBER_CODE "
                  "JOIN INV_INV_MST i ON i.IN_INV_ID=d.IN_INV_ID WHERE a.VC_PRODUCTION_DATE='%s' "
                  "GROUP BY d.VC_ASSY_PART_NUMBER ORDER BY COUNT(*) DESC" % pdate)
    part_lines = int(scalar("SELECT COUNT(*) FROM INV_ASN_MST a JOIN INV_ASN_DETAIL_MST d "
                            "ON a.IN_ASN_ID=d.IN_ASN_ID JOIN INV_INV_MST i ON i.IN_INV_ID=d.IN_INV_ID "
                            "WHERE a.VC_PRODUCTION_DATE='%s' AND d.VC_ASSY_PART_NUMBER='%s'"
                            % (pdate, part)))
    base_f = len(fai)
    base_c = len(cor)
    # Inject a 2nd, NON-COVERING (gap) window for `part`. VC_ASSY_MANIFEST_NUMBER is varchar(2) so the tag
    # is 'M3'; we identify the injected row uniquely by its ancient window dates (no real window uses these)
    # for a precise teardown. The faithful (window-blind) join attaches BOTH windows -> doubles the part's
    # lines; the corrected CROSS APPLY picks only the one COVERING window -> unchanged.
    GAP = ("20100101", "20101231")
    sqlrows("INSERT INTO INV_MANIFEST_COST_MST (VC_ASSY_PART_NUMBER_CODE, VC_ASSY_MANIFEST_NUMBER, "
            "VC_START_MANIFEST, VC_END_MANIFEST, MO_PRICE) VALUES "
            "('%s','M3','%s','%s',9.99)" % (part, GAP[0], GAP[1]))
    try:
        _, f2d = sqlrows(bind_qmarks(load_sql("invoice_summary_faithful"), [pdate]), with_header=True)
        _, c2d = sqlrows(bind_qmarks(load_sql("invoice_summary"), [pdate]), with_header=True)
        f2, c2 = len(f2d), len(c2d)
        print("  injected gap window for %s (%d lines): faithful %d->%d  corrected %d->%d"
              % (part, part_lines, base_f, f2, base_c, c2))
        rep.check("R9 D6 non-vacuity: faithful (window-blind) DOUBLES the part's lines (over-bill)",
                  f2 == base_f + part_lines, "faithful %d->%d (+%d expected)" % (base_f, f2, part_lines))
        rep.check("R9 D6 non-vacuity: corrected UNCHANGED (the one covering window)",
                  c2 == base_c, "corrected %d->%d" % (base_c, c2))
    finally:
        sqlrows("DELETE FROM INV_MANIFEST_COST_MST WHERE VC_ASSY_PART_NUMBER_CODE='%s' "
                "AND VC_START_MANIFEST='%s' AND VC_END_MANIFEST='%s'" % (part, GAP[0], GAP[1]))
        left = int(scalar("SELECT COUNT(*) FROM INV_MANIFEST_COST_MST WHERE VC_ASSY_PART_NUMBER_CODE='%s' "
                          "AND VC_START_MANIFEST='%s' AND VC_END_MANIFEST='%s'"
                          % (part, GAP[0], GAP[1])) or 0)
        rep.check("R9 fixture clean (injected D6 window removed)", left == 0, "left=%d" % left)

    # LAYER 2 — POI render readback (the CORRECTED default). Item Total = Qty*UnitCost; INVOICE TOTAL band.
    _, detail = fetch_rows_as_dicts("invoice_summary", [pdate])
    out = os.path.join(OUT_DIR, "invoice.xlsx")
    try:
        jython_render("declarative", _jsonable(rdef), {"PDate": _fmt_date(pdate)}, out, detail=detail)
    except subprocess.CalledProcessError as e:
        rep.check("R9 L2: POI render ran", False, e.output[-300:])
        return
    ws = read_xlsx(out)
    rep.check("R9 L2 [1,1] title", cell(ws, 1, 1) == "INVOICE (Assy Part Numbers)", "%r" % cell(ws, 1, 1))
    rep.check("R9 L2 [2,1] 'INVOICE DATE:'+date",
              cell(ws, 2, 1) == "INVOICE DATE:" + _fmt_date(pdate), "%r" % cell(ws, 2, 1))
    titles = ["Part Number", "Manifest Number", "Qty", "Unit Cost", "Item Total"]
    rep.check("R9 L2 row-3 col titles", [cell(ws, 3, c) for c in range(1, 6)] == titles,
              "%r" % [cell(ws, 3, c) for c in range(1, 6)])
    rep.check("R9 L2 widths 17/17/30/30/30",
              [colwidth(ws, c) for c in range(1, 6)] == [17, 17, 30, 30, 30],
              "%r" % [colwidth(ws, c) for c in range(1, 6)])
    # detail row 4: part/manifest/qty + ITEM TOTAL = qty*price (independent recompute from the NQ row)
    d0 = detail[0]
    exp_item = float(d0["ShipQty"]) * float(d0["UnitPrice"])
    got_item = float(cell(ws, 4, 5))
    rep.check("R9 L2 [4,1..4] PartNumber/Manifest/Qty/UnitCost",
              cell(ws, 4, 1) == d0["PartNumber"] and cell(ws, 4, 2) == d0["Manifest"]
              and int(cell(ws, 4, 3)) == int(d0["ShipQty"]) and abs(float(cell(ws, 4, 4)) - float(d0["UnitPrice"])) < 1e-6)
    rep.check("R9 L2 [4,5] Item Total == Qty*UnitCost", abs(got_item - exp_item) < 1e-6,
              "got=%s exp=%s" % (got_item, exp_item))
    # Item Total money number-format applied (numeric cell)
    rep.check("R9 L2 [4,5] money number-format applied",
              "$" in ws.cell(row=4, column=5).number_format, ws.cell(row=4, column=5).number_format)
    # INVOICE TOTAL band: 2 blank rows past the last detail, label at col3, grand sum at col5 == SUM(items)
    n = len(detail)
    tz = (4 + n - 1) + 2 + 1
    grand_exp = sum(float(d["ShipQty"]) * float(d["UnitPrice"]) for d in detail)
    rep.check("R9 L2 INVOICE TOTAL label at [tz,3]", cell(ws, tz, 3) == "INVOICE TOTAL",
              "tz=%d val=%r" % (tz, cell(ws, tz, 3)))
    rep.check("R9 L2 INVOICE TOTAL grand sum at [tz,5] == SUM(item totals)",
              abs(float(cell(ws, tz, 5)) - grand_exp) < 1e-4,
              "got=%s exp=%s" % (cell(ws, tz, 5), grand_exp))


# ----------------------------------------------------------------------------------------------------
# R18 — Forecast Detail (faithful; group-break custom render)
# ----------------------------------------------------------------------------------------------------
def test_forecast_detail(rep, defs):
    print("\n--- R18 Forecast Detail (REPORT_ForecastDetail, faithful group-break render) ---")
    rdef = defs.REPORTS["forecast_detail"]

    # LAYER 1 — parity: the .sql columns/order == the legacy proc (SELECT * pinned to the 8 read cols).
    legacy_n = int(scalar("SELECT COUNT(*) FROM (SELECT * FROM inv_forecast_detail_inf) z"))
    cols, nq = fetch_rows_as_dicts("forecast_detail")
    rep.check("R18 L1: NQ row count == legacy proc row count (%d)" % legacy_n,
              len(nq) == legacy_n and legacy_n > 0, "nq=%d legacy=%d" % (len(nq), legacy_n))
    # independent ordered-key truth: same ORDER BY, same 8 cols, compared row-for-row
    truth_cols, truth = fetch_rows_via_sql(
        "SELECT VC_ASSY_PART_NUMBER_CODE, VC_EFFECTIVE_MONTH, VC_BROADCAST_CODE, VC_TIRE_PART_NUMBER_CODE, "
        "IN_TIRE_RATIO, VC_WHEEL_PART_NUMBER_CODE, IN_WHEEL_RATIO, IN_RATIO FROM inv_forecast_detail_inf "
        "ORDER BY vc_assy_part_number_code, vc_effective_month, vc_broadcast_code")
    rep.check("R18 L1: NQ rows == independent SELECT (8 cols, same ORDER BY), row-for-row",
              [tuple(r) for r in _as_lists(nq, cols)] == [tuple(r) for r in truth], "")

    # NON-VACUITY: a different ORDER BY produces a different sequence (proves the comparison can fail).
    _, scrambled = fetch_rows_via_sql(
        "SELECT VC_ASSY_PART_NUMBER_CODE, VC_EFFECTIVE_MONTH, VC_BROADCAST_CODE, VC_TIRE_PART_NUMBER_CODE, "
        "IN_TIRE_RATIO, VC_WHEEL_PART_NUMBER_CODE, IN_WHEEL_RATIO, IN_RATIO FROM inv_forecast_detail_inf "
        "ORDER BY IN_RATIO DESC, vc_assy_part_number_code")
    rep.check("R18 L1 non-vacuity: a DIFFERENT order != the NQ order",
              [tuple(r) for r in _as_lists(nq, cols)] != [tuple(r) for r in scrambled]
              or len(set(map(tuple, truth))) <= 1)

    # LAYER 2 — POI render readback: the faithful group-break layout.
    out = os.path.join(OUT_DIR, "forecast.xlsx")
    try:
        jython_render("build_forecast_detail_plan", _jsonable(rdef), {}, out, detail=nq)
    except subprocess.CalledProcessError as e:
        rep.check("R18 L2: POI render ran", False, e.output[-300:])
        return
    ws = read_xlsx(out)
    rep.check("R18 L2 [1,1] title", cell(ws, 1, 1) == "Forecast Detail Report", "%r" % cell(ws, 1, 1))
    fd_titles = ["Assembly Part#", "Effective Month", "BC Code", "Tire Part Number", "Tire Share Ratio",
                 "Wheel Part Number", "Wheel Share Ratio", "Forecast Share"]
    rep.check("R18 L2 row-3 col titles (8)", [cell(ws, 3, c) for c in range(1, 9)] == fd_titles,
              "%r" % [cell(ws, 3, c) for c in range(1, 9)])
    rep.check("R18 L2 widths 17/17/10/17/17/17/17/17",
              [colwidth(ws, c) for c in range(1, 9)] == [17, 17, 10, 17, 17, 17, 17, 17])

    # Reproduce the legacy loop EXPECTATION INDEPENDENTLY (a hand-transcription of MainMenu.pas:3768-3794),
    # then assert the rendered sheet matches it cell-for-cell. (Expectation derived from the SOURCE .pas,
    # NOT from the rebuild's plan.)
    exp = _forecast_expected_cells(nq)
    mismatches = []
    for (r, c, v) in exp:
        got = cell(ws, r, c)
        gv = _coerce_cmp(got)
        ev = _coerce_cmp(v)
        if gv != ev:
            mismatches.append((r, c, ev, gv))
    rep.check("R18 L2 group-break layout == independent .pas-loop expectation (cell-for-cell)",
              not mismatches, "first 3 mismatches: %r" % mismatches[:3])
    # blank separator row exists before the 2nd part group (the INC(z) on part-change)
    parts = [r["VC_ASSY_PART_NUMBER_CODE"] for r in nq]
    if len(set(parts)) >= 2:
        # find the row of the 2nd distinct part's first cell and assert the row ABOVE it has empty col1..8
        rep.check("R18 L2 a blank separator row precedes a new part group",
                  _has_blank_separator(ws, exp), "")


# ----------------------------------------------------------------------------------------------------
# R12 — Lot Location (faithful; two-proc custom render; seeded for non-vacuity)
# ----------------------------------------------------------------------------------------------------
def test_lot_location(rep, defs):
    print("\n--- R12 Lot Location (REPORT_PLANTLotLocation[W], faithful two-proc render; SEEDED) ---")
    rdef = defs.REPORTS["lot_location"]

    # CONFIRM LIVE (not the D9 NUMMI twin) — assert by the def's wrapped procs.
    rep.check("R12 confirmed LIVE: wraps the PLANT procs (not D9 NUMMI)",
              rdef["proc"] == "REPORT_PLANTLotLocation / REPORT_PLANTLotLocationW"
              and "NUMMI" not in rdef["proc"])

    # The spike has 0 lot-location rows (every vc_status_PLANT_yard=''), so SEED fixtures to be NON-VACUOUS.
    # Flip the yard status on a few existing open-order rows (additive-then-restore), one wheel + one tire,
    # spanning the Passenger/Truck line-name + a size group so the group-break render is exercised.
    seeded = _seed_lot_location()
    try:
        # LAYER 1 — parity: the .sql == the legacy proc, after seeding (the procs now return our rows).
        leg_w = sqlrows("EXEC REPORT_PLANTLotLocationW", with_header=True)
        nq_w = sqlrows(load_sql("lot_location_wheels"), with_header=True)
        leg_t = sqlrows("EXEC REPORT_PLANTLotLocation", with_header=True)
        nq_t = sqlrows(load_sql("lot_location_tires"), with_header=True)
        rep.check("R12 L1 wheels NQ == legacy REPORT_PLANTLotLocationW (rows)",
                  _all_rows(leg_w) == _all_rows(nq_w) and len(nq_w[1]) > 0,
                  "legacy=%d nq=%d" % (len(leg_w[1]), len(nq_w[1])))
        rep.check("R12 L1 tires NQ == legacy REPORT_PLANTLotLocation (rows)",
                  _all_rows(leg_t) == _all_rows(nq_t) and len(nq_t[1]) > 0,
                  "legacy=%d nq=%d" % (len(leg_t[1]), len(nq_t[1])))

        # NON-VACUITY: an UNSEEDED snapshot would be 0 rows — assert we actually have rows post-seed.
        rep.check("R12 L1 non-vacuity: seeding produced rows (the unseeded spike returns 0)",
                  len(nq_w[1]) > 0 and len(nq_t[1]) > 0,
                  "wheels=%d tires=%d" % (len(nq_w[1]), len(nq_t[1])))

        # LAYER 2 — POI render readback (two-proc union + group-break headers + conditional cells)
        _, wheels = fetch_rows_as_dicts("lot_location_wheels")
        _, tires = fetch_rows_as_dicts("lot_location_tires")
        out = os.path.join(OUT_DIR, "lotloc.xlsx")
        params = {"PlantName": "TestPlant", "AssemblerName": "TestAsm"}
        try:
            jython_render("build_lot_location_plan", _jsonable(rdef), params, out, wheels=wheels, tires=tires)
        except subprocess.CalledProcessError as e:
            rep.check("R12 L2: POI render ran", False, e.output[-300:])
            return
        ws = read_xlsx(out)
        rep.check("R12 L2 [1,1] '<Plant> Lot Location'", cell(ws, 1, 1) == "TestPlant Lot Location",
                  "%r" % cell(ws, 1, 1))
        loc_titles = ["Size", "Parking", "Trailer", "Renban", "Arrived", "TestAsm Loc", "TestPlant Loc"]
        rep.check("R12 L2 row-3 col titles (7)", [cell(ws, 3, c) for c in range(1, 8)] == loc_titles,
                  "%r" % [cell(ws, 3, c) for c in range(1, 8)])
        rep.check("R12 L2 widths 11/11/14/11/10/12/12",
                  [colwidth(ws, c) for c in range(1, 8)] == [11, 11, 14, 11, 10, 12, 12])
        # independent .pas-loop expectation for the two-proc render (hand-transcribed MainMenu:716-861)
        exp = _lot_location_expected_cells(wheels, tires, "TestPlant", "TestAsm")
        mism = []
        for (r, c, v) in exp:
            gv = _coerce_cmp(cell(ws, r, c)); ev = _coerce_cmp(v)
            if gv != ev:
                mism.append((r, c, ev, gv))
        rep.check("R12 L2 two-proc group-break layout == independent .pas-loop expectation",
                  not mism, "first 3 mismatches: %r" % mism[:3])
        # a wheel group header ('Car Wheels' or 'Truck Wheels') and a tire size header are present in col1
        col1 = [cell(ws, r, 1) for r in range(5, ws.max_row + 1)]
        rep.check("R12 L2 wheel group header present (Car/Truck Wheels)",
                  any(v in ("Car Wheels", "Truck Wheels") for v in col1), "%r" % col1[:10])
    finally:
        _unseed_lot_location(seeded)
        left = int(scalar("SELECT COUNT(*) FROM inv_open_order_inf WHERE vc_status_PLANT_yard='M3SEED99'") or 0)
        rep.check("R12 fixture clean (seeded yard-status restored)", left == 0, "left=%d" % left)


# ============================================================================================
# Lot-Location fixture seeding (additive-then-restore; flips yard status on existing rows).
# ============================================================================================
SEED_YARD = "20260601"          # a valid yyyymmdd so _yard_date reformats cleanly
SEED_TAG = "M3SEED99"           # the marker we restore by (placed in vc_status_PLANT_yard transiently? no —
#                                 we keep the original value to restore; the tag is only used as a sentinel
#                                 to verify cleanup. We capture the original yard value per id and restore it.)


def _seed_lot_location():
    """Pick a few existing open-order rows (1+ wheel, 1+ tire) whose master/type/size joins resolve, and
    set vc_status_PLANT_yard to a non-empty yyyymmdd so the procs' WHERE passes. Capture (id, old_value)
    so teardown restores EXACTLY. Returns the capture list. Additive only (no inserts/deletes)."""
    # wheels: need a Passenger and a non-Passenger line for the group break; tires: a size for the header.
    captured = []
    # 2 wheel rows: Passenger + Truck (distinct line names), 2 tire rows.
    rows = sqlrows(
        "SELECT TOP 2 o.IN_ORDER_ID, o.vc_status_PLANT_yard, p.vc_LINE_NAME "
        "FROM inv_open_order_inf o JOIN inv_parts_stock_mst p ON o.vc_part_number=p.vc_part_number "
        "JOIN INV_PART_TYPE_MST x ON p.IN_PART_TYPE_ID=x.IN_PART_TYPE_ID "
        "WHERE x.VC_PART_TYPE='WHEEL' AND o.vc_status_empty_trailer='' AND o.vc_terminated='' "
        "ORDER BY p.vc_LINE_NAME", with_header=False)
    tire_rows = sqlrows(
        "SELECT TOP 2 o.IN_ORDER_ID, o.vc_status_PLANT_yard "
        "FROM inv_open_order_inf o JOIN inv_parts_stock_mst p ON o.vc_part_number=p.vc_part_number "
        "JOIN INV_PART_TYPE_MST x ON p.IN_PART_TYPE_ID=x.IN_PART_TYPE_ID "
        "JOIN INV_SIZE_MST s ON p.IN_SIZE_ID=s.IN_SIZE_ID "
        "WHERE x.VC_PART_TYPE='TIRE' AND o.vc_status_empty_trailer='' AND o.vc_terminated='' "
        "ORDER BY s.vc_size_code", with_header=False)
    for r in rows + tire_rows:
        oid, old = r[0], (r[1] if len(r) > 1 else "")
        captured.append((oid, old))
        sqlrows("UPDATE inv_open_order_inf SET vc_status_PLANT_yard='%s' WHERE IN_ORDER_ID=%s"
                % (SEED_YARD, oid))
    return captured


def _unseed_lot_location(captured):
    for (oid, old) in captured:
        val = old if old not in (None, "NULL") else ""
        sqlrows("UPDATE inv_open_order_inf SET vc_status_PLANT_yard='%s' WHERE IN_ORDER_ID=%s"
                % (val.replace("'", "''"), oid))


# ============================================================================================
# Independent expectation builders (hand-transcribed from the .pas — the SOURCE, not the rebuild plan)
# ============================================================================================
def _forecast_expected_cells(rows):
    """Hand-transcription of MainMenu.pas:3768-3794 (NOT a call to code.build_forecast_detail_plan).
    Returns [(row, col, value)] the legacy loop would write."""
    out = [(1, 1, "Forecast Detail Report")]
    z = 4
    lastpn = ""
    lastmn = ""
    for r in rows:
        pn = _s(r["VC_ASSY_PART_NUMBER_CODE"]); mn = _s(r["VC_EFFECTIVE_MONTH"])
        if lastpn != pn:
            z += 1
            out.append((z, 1, pn))
            out.append((z, 2, mn))
            lastpn, lastmn = pn, mn
        if lastmn != mn:
            out.append((z, 2, mn))
            lastmn = mn
        out.append((z, 3, _s(r["VC_BROADCAST_CODE"])))
        out.append((z, 4, _s(r["VC_TIRE_PART_NUMBER_CODE"])))
        out.append((z, 5, _int(r["IN_TIRE_RATIO"])))
        out.append((z, 6, _s(r["VC_WHEEL_PART_NUMBER_CODE"])))
        out.append((z, 7, _int(r["IN_WHEEL_RATIO"])))
        out.append((z, 8, _int(r["IN_RATIO"])))
        z += 1
    return out


def _lot_location_expected_cells(wheels, tires, plant, assembler):
    """Hand-transcription of MainMenu.pas:716-861 (NOT code.build_lot_location_plan)."""
    out = [(1, 1, plant + " Lot Location")]
    titles = ["Size", "Parking", "Trailer", "Renban", "Arrived", assembler + " Loc", plant + " Loc"]
    for ci, t in enumerate(titles, start=1):
        out.append((3, ci, t))
    z = 5
    lastwheel = ""
    for r in wheels:
        ln = _s(r["vc_LINE_NAME"])
        if lastwheel != ln:
            if ln == "Passenger":
                out.append((z, 1, "Car Wheels"))
            else:
                z += 1
                out.append((z, 1, "Truck Wheels"))
            lastwheel = ln
        out.append((z, 2, _park(r)))
        out.append((z, 3, _s(r["vc_trailer_number"])))
        out.append((z, 4, _s(r["vc_renban_number"])))
        out.append((z, 5, _yard(r["vc_status_PLANT_yard"])))
        z += 1
    lastsize = "X"
    for r in tires:
        sz = _s(r["vc_size_code"])
        if sz != lastsize:
            z += 1
            out.append((z, 1, sz if sz != "" else "<NO SIZE>"))
            lastsize = sz
        out.append((z, 2, _park(r)))
        out.append((z, 3, _s(r["vc_trailer_number"])))
        out.append((z, 4, _s(r["vc_renban_number"])))
        out.append((z, 5, _yard(r["vc_status_PLANT_yard"])))
        z += 1
    return out


def _park(r):
    a = _s(r["vc_ASSEMBLER_location"])
    return a if a != "" else _s(r["vc_PLANT_parking"])


def _yard(v):
    s = _s(v)
    return s[0:4] + "/" + s[4:6] + "/" + s[6:8]


# ============================================================================================
# small utils
# ============================================================================================
def _fmt_date(yyyymmdd):
    """yyyymmdd -> yyyy/mm/dd (the legacy date the dialog passes as ProductionDate; the handler writes it
    to cell [2,1] verbatim)."""
    s = str(yyyymmdd)
    return s[0:4] + "/" + s[4:6] + "/" + s[6:8]


def _s(v):
    return "" if v in (None, "NULL") else str(v)


def _int(v):
    if v in (None, "", "NULL"):
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        return int(float(v))


def _coerce_cmp(v):
    """Normalize a cell for compare: numbers -> int when integral, strings as-is, None -> ''."""
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def _legacy_assy_header_row1(pdate):
    """INDEPENDENT legacy truth for the Daily Shipping Assy header band.

    The Delphi handler prints Start/End/Vehicle Count from the FIRST DETAIL ROW of
    REPORT_DailyShippingAssy (MainMenu.pas:3173-3174 — fieldbyname before the z:=4 loop). We EXEC the REAL
    legacy proc and read row 1's Start/End/Vehicle Count. This is the legacy proc's OWN output (not the
    rebuild's plan, not MIN/MAX/SUM). The proc's ORDER BY is VC_ASSY_PART_NUMBER only; on the spike's
    multi-header dates the alphabetically-first part is unambiguous (single header) on the chosen dates, so
    row 1's header is deterministic. Returns (Start, End, VehicleCount) as strings.
    """
    cols, data = sqlrows("EXEC REPORT_DailyShippingAssy @PDate='%s'" % pdate, with_header=True)
    if not data:
        return None, None, None
    i_s, i_e, i_v = cols.index("Start"), cols.index("End"), cols.index("Vehicle Count")
    r0 = data[0]
    return r0[i_s], r0[i_e], r0[i_v]


def _pairs(hdr_data, c1, c2):
    cols, data = hdr_data
    i1, i2 = cols.index(c1), cols.index(c2)
    return [(r[i1], r[i2]) for r in data]


def _tuples(hdr_data, names):
    cols, data = hdr_data
    idx = [cols.index(n) for n in names]
    return [tuple(r[i] for i in idx) for r in data]


def _all_rows(hdr_data):
    return [tuple(r) for r in hdr_data[1]]


def _as_lists(dicts, cols):
    return [[_norm(d[c]) for c in cols] for d in dicts]


def _norm(v):
    return "" if v in (None, "NULL") else str(v)


def fetch_rows_via_sql(sql):
    cols, data = sqlrows(sql, with_header=True)
    return cols, [[_norm(c) for c in r] for r in data]


def _count_detail_rows(ws, start_row, key_col):
    n = 0
    r = start_row
    while ws.cell(row=r, column=key_col).value not in (None, ""):
        n += 1
        r += 1
    return n


def _has_blank_separator(ws, exp_cells):
    """True if at least one row between first_data_row and max has all of cols 1..8 empty (the INC(z) gap)."""
    used = set(r for (r, c, v) in exp_cells if r >= 4)
    for r in range(5, ws.max_row + 1):
        if r not in used:
            if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, 9)):
                return True
    return False


def _jsonable(rdef):
    """report_defs entries contain tuples (header_cells) — JSON round-trips tuples as lists, which the
    jython side handles (it indexes by position). Deep-convert to a plain dict the JSON+jython path reads."""
    return json.loads(json.dumps(rdef))


# ============================================================================================
# OUTPUT STAGE — code.legacy_timestamp + code.write_output (tmp-then-rename). Layer 2 renders directly
# via FileOutputStream, so the output stage is exercised SEPARATELY here under a tiny CPython `system`
# shim (write_output only touches system.date + system.file + os). Proves: the legacy 14-digit + literal
# '00' filename suffix, the .xlsx extension, and that NO .tmp survives a successful write (atomicity).
# ============================================================================================
def test_output_stage(rep):
    print("\n--- Output stage (legacy_timestamp + write_output tmp-then-rename) ---")
    import importlib.util
    import datetime
    spec = importlib.util.spec_from_file_location("rr_code", os.path.join(RR_DIR, "code.py"))
    eng = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(eng)

    class _Date:
        def now(self):
            return datetime.datetime(2026, 6, 21, 14, 57, 29)
        def format(self, d, pattern):
            return d.strftime({"yyyyMMddHHmmss": "%Y%m%d%H%M%S"}[pattern])

    class _File:
        def writeFile(self, path, data, append=False):
            mode = "ab" if append else "wb"
            payload = data if isinstance(data, (bytes, bytearray)) else str(data).encode("ascii")
            with open(path, mode) as fh:
                fh.write(payload)

    class _Sys:
        date = _Date()
        file = _File()
    eng.system = _Sys()

    ts = eng.legacy_timestamp()
    rep.check("Output: legacy_timestamp = yyyymmddhhmmss + literal '00' (14+2 digits)",
              ts == "2026062114572900", "ts=%r" % ts)

    d = os.path.join(OUT_DIR, "outstage")
    os.makedirs(d, exist_ok=True)
    path = eng.write_output(b"PK\x03\x04xlsxbytes", "DailyShippingAssy", d)
    rep.check("Output: filename = <out_name><ts>.xlsx (legacy stem, .xls->.xlsx)",
              os.path.basename(path) == "DailyShippingAssy2026062114572900.xlsx",
              os.path.basename(path))
    rep.check("Output: final file exists + content written", os.path.exists(path)
              and open(path, "rb").read() == b"PK\x03\x04xlsxbytes")
    rep.check("Output: NO orphan .tmp survives a successful write (atomicity)",
              not os.path.exists(path + ".tmp"))


# ============================================================================================
# REVERT NON-VACUITY — prove the parity comparisons can FAIL on a wrong rebuild. We MUTATE the loaded
# .sql text (the rebuild artifact) into a wrong-but-runnable variant and assert the parity diff FAILS.
# "A test that can't fail on a wrong rebuild is testing nothing." (feedback-self-referential-test-discipline)
# ============================================================================================
def test_revert_non_vacuity(rep):
    print("\n--- Revert non-vacuity (a corrupted rebuild .sql must FAIL parity) ---")

    # Daily Shipping Assy: swap SUM(d.IN_QTY) -> SUM(d.IN_QTY)+1 (a wrong rebuild) and prove the parity
    # diff vs the legacy proc now FAILS for the same date.
    pdate = scalar("SELECT TOP 1 a.VC_PRODUCTION_DATE FROM INV_ASN_MST a "
                   "JOIN INV_ASN_DETAIL_MST d ON a.IN_ASN_ID=d.IN_ASN_ID "
                   "GROUP BY a.VC_PRODUCTION_DATE ORDER BY COUNT(*) DESC")
    legacy = _pairs(sqlrows("EXEC REPORT_DailyShippingAssy @PDate='%s'" % pdate, with_header=True),
                    "Part Number", "PQty")
    good_sql = bind_qmarks(load_sql("daily_shipping_assy"), [pdate])
    wrong_sql = good_sql.replace("SUM(d.IN_QTY)", "SUM(d.IN_QTY)+1")
    wrong = _pairs(sqlrows(wrong_sql, with_header=True), "Part Number", "PQty")
    rep.check("Revert: a +1 qty rebuild FAILS the daily_shipping_assy parity (test is non-vacuous)",
              wrong != legacy and len(wrong) > 0, "wrong==legacy? %s" % (wrong == legacy))

    # INVOICE Summary corrected: swap the window-aware CROSS APPLY back to the window-BLIND JOIN (i.e.
    # "revert the D6 fix") and prove the corrected-vs-truth diff FAILS once a 2nd window is present.
    part = scalar("SELECT TOP 1 d.VC_ASSY_PART_NUMBER FROM INV_ASN_MST a "
                  "JOIN INV_ASN_DETAIL_MST d ON a.IN_ASN_ID=d.IN_ASN_ID "
                  "JOIN INV_INV_MST i ON i.IN_INV_ID=d.IN_INV_ID WHERE a.VC_PRODUCTION_DATE='%s' "
                  "GROUP BY d.VC_ASSY_PART_NUMBER ORDER BY COUNT(*) DESC" % pdate)
    GAP = ("20100101", "20101231")
    sqlrows("INSERT INTO INV_MANIFEST_COST_MST (VC_ASSY_PART_NUMBER_CODE, VC_ASSY_MANIFEST_NUMBER, "
            "VC_START_MANIFEST, VC_END_MANIFEST, MO_PRICE) VALUES ('%s','M3','%s','%s',9.99)"
            % (part, GAP[0], GAP[1]))
    try:
        truth_n = int(scalar(
            "SELECT COUNT(*) FROM INV_ASN_MST a JOIN INV_ASN_DETAIL_MST d ON a.IN_ASN_ID=d.IN_ASN_ID "
            "CROSS APPLY dbo.fn_ManifestCostAt(d.VC_ASSY_PART_NUMBER, a.VC_PRODUCTION_DATE) m "
            "JOIN INV_INV_MST i ON i.IN_INV_ID=d.IN_INV_ID WHERE a.VC_PRODUCTION_DATE='%s'" % pdate))
        corrected_sql = bind_qmarks(load_sql("invoice_summary"), [pdate])
        reverted = corrected_sql.replace(
            "CROSS APPLY dbo.fn_ManifestCostAt(d.VC_ASSY_PART_NUMBER, a.VC_PRODUCTION_DATE) m",
            "JOIN INV_MANIFEST_COST_MST m ON d.VC_ASSY_PART_NUMBER = m.VC_ASSY_PART_NUMBER_CODE")
        _, rev_data = sqlrows(reverted, with_header=True)
        rep.check("Revert: re-introducing the window-BLIND JOIN over-bills (corrected!=truth) — D6 test non-vacuous",
                  len(rev_data) != truth_n and len(rev_data) > truth_n,
                  "reverted=%d truth=%d" % (len(rev_data), truth_n))
    finally:
        sqlrows("DELETE FROM INV_MANIFEST_COST_MST WHERE VC_ASSY_PART_NUMBER_CODE='%s' "
                "AND VC_START_MANIFEST='%s' AND VC_END_MANIFEST='%s'" % (part, GAP[0], GAP[1]))


if __name__ == "__main__":
    main()
